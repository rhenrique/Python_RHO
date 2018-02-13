# encoding: utf-8
# encoding: iso-8859-1
# encoding: win-1252
## COMPILAR: pyinstaller.exe --onefile --icon=ORC_icon.ico ORCAM_check.py
import pyodbc
import sys
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
import shutil
import subprocess
import time
import os
from distutils.core import setup
from ConfigParser import SafeConfigParser
import glob
import socket
import os.path
from openpyxl import Workbook
from ctypes import *

wb = Workbook()
ws1 = wb.active
ws1.title = 'Aco'
ws2 = wb.create_sheet('Alum')
ws1['A1'] = 'Conta'
ws1['B1'] = 'Valor (ac. do mes)'
ws2['A1'] = 'Conta'
ws2['B1'] = 'Valor (ac. do mes)'


path_aco = r'F:\FTP\Orcam\128\ORCAM.txt'
path_alum = r'F:\FTP\Orcam\130\ORCAM.txt'
print('ACO - ultima vez gerado: %s' % time.ctime(os.path.getmtime(path_aco)))
print('ALUM - ultima vez gerado: %s' % time.ctime(os.path.getmtime(path_alum)))
aco = time.ctime(os.path.getmtime(path_aco))
alum = time.ctime(os.path.getmtime(path_alum))

cdate = time.strftime("%Y%m")
ins_date = time.strftime("%Y%m%d")
emailfrom = "ORCAM_report@maxionwheels.com"
server = 'LIM-SQL12P1'
username = 'dsserver'
password = 'Maxion123@'
driver = '{SQL Server}' # Driver necessário para conecta ao banco de dados do MSSQL
port = '1433'
emailto = 'rafael.oliveira@maxionwheels.com;alex.bacalhau@maxionwheels.com;marcos.degaspre@maxionwheels.com;gilberto.ferreira@maxionwheels.com'

def conn(ambiente):
	if ambiente == "aco":
		database = 'PROD_ORCAM_LMS'
		cnn = pyodbc.connect('DRIVER='+driver+';PORT=port;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+
					';PWD='+password)
		cnn.setencoding(str, encoding='utf-8')
		cnn.setencoding(unicode, encoding='utf-8', ctype=pyodbc.SQL_CHAR)
		cursor = cnn.cursor()
		app(cnn, cursor)
	if ambiente == "alum":
		database = 'PROD_ORCAM_LMA'
		cnn = pyodbc.connect('DRIVER='+driver+';PORT=port;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+
					';PWD='+password)
		cnn.setencoding(str, encoding='utf-8')
		cnn.setencoding(unicode, encoding='utf-8', ctype=pyodbc.SQL_CHAR)
		cursor = cnn.cursor()
		msgfinal2, body_alum = ambAlum(cnn, cursor)
		return msgfinal2, body_alum

def ambAlum(cnn, cursor):
	row = cursor.execute("select * from Grupos")
	i = 2
	body4 = []
	for row in cursor.fetchall():
		for conta in cursor.execute("select sum(Vr_real) from orcam where anomes=? and grupo=?", cdate, row.Grupo):
			if conta[0] != None:
				
				value_real = str("%0.2f" % (conta[0])).replace('.',',')
				ws2.cell(row=i, column=1).value = row.Grupo
				ws2.cell(row=i, column=2).value = value_real
				body4.append(u"""<html> <head> 
			<style> 
			table {
				font-family: arial, sans-serif;
				border-collapse: separate;
				width: 20%
				}
			td, th {
				border: 2px solid #dddddd;
				text-align: left;
				padding: 5px;
				}

			tr:nth-child(even) {
				background-color: #dddddd;
				}
</style>
			 </head><body><table style="display: inline-block; border: 1px solid; ">
			 <tr> <td> """ + str(row.Grupo) + u"""</td> 
			 <td>""" + str(conta[0]) + u"""</td></tr></table></body></html>""")
				i = i + 1
		msgfinal2 = ''.join(body4)
	if body4:
		body_alum = """<html>
<head>
<style>
table {
    font-family: arial, sans-serif;
    border-collapse: separate;
    width: 20%;
}

td, th {
    border: 2px solid #dddddd;
    text-align: left;
    padding: 5px;
}

tr:nth-child(even) {
    background-color: #dddddd;
}
</style>
</head>
<body>
	<table style="display: inline-block; border: 1px solid; ">
		<tr>
		<th> ALUM - Conta</th>
		<th> Valor Real (Soma)</th>
		</tr>
	</table>
</body>
</html>
"""
	return msgfinal2, body_alum
			 
		
		
def app(cnn, cursor):
	row = cursor.execute("select * from Grupos")
	i = 2
	body = []
	for row in cursor.fetchall():
		for conta in cursor.execute("select sum(Vr_real) from orcam where anomes=? and grupo=?", cdate, row.Grupo):
			sub = u"ORÇAM: Verifica Atualização (INTERFACE - QAD)"
			if conta[0] != None:
				value_real = str("%0.2f" % (conta[0])).replace('.',',')
				ws1.cell(row=i, column=1).value = row.Grupo
				ws1.cell(row=i, column=2).value = value_real
				body.append(u"""<html> <head> 
			<style> 
			table {
				font-family: arial, sans-serif;
				border-collapse: separate;
				width: 20%
				}
			td, th {
				border: 2px solid #dddddd;
				text-align: left;
				padding: 5px;
				}

			tr:nth-child(even) {
				background-color: #dddddd;
				}
</style>
			 </head><body><table style="display: inline-block; border: 1px solid; float: left; ">
			 <tr> <td> """ + str(row.Grupo) + u"""</td> 
			 <td>""" + str(conta[0]) + u"""</td></tr></table></body></html>""")
				i = i + 1
			
	if body:
		body2 = u"""<html>
<head>
<style>
table {
    font-family: arial, sans-serif;
    border-collapse: separate;
    width: 20%;
}

td, th {
    border: 2px solid #dddddd;
    text-align: left;
    padding: 5px;
}

tr:nth-child(even) {
    background-color: #dddddd;
}
</style>
</head>
<body>
	<table style="display: inline-block; border: 1px solid; float: left; ">
		<tr>
		<th> AÇO - Conta</th>
		<th> Valor Real (Soma)</th>
		</tr>
	</table>
</body>
</html>
"""
		emailfrom = "ORCAM_report@maxionwheels.com"
		body3 = u"""<html>
		<head>
		<font face="Calibri" size="3" color = "black">
		</head>
		<body>
			<b>Última vez em que a interface do QAD foi executada:</b> 
			<br>
			<b>- AÇO: </b> """ + aco + u"""
			<br>
			<b>- ALUM: </b> """ + alum + u"""
			<br>
			<br>
			<b>Explicação:</b> a interface do QAD é executado todos os dias nos seguintes horários:<br>
			- 11h30min;
			<br>
			- 23h35min,
			
			<br>
			<br>
			<b>Verificar sempre a data em que a última vez foi gerado e comparar com a data atual.</b> 
			<br>
			<br>
			
		
		
		</body>
		</font>
		</html>"""
		body5, body_alum = conn('alum')
		fl = 'ORCAM_' + cdate + '.xlsx'
		print fl
		wb.save(fl)
		msgfinal = ''.join(body)
		msglala = body3
		msg = MIMEMultipart()
		msg.attach(MIMEText(msglala, 'html', 'utf-8'))
		msg["From"] = emailfrom
		msg["Subject"] = sub
		msg["To"] = emailto
		ctype, encoding = mimetypes.guess_type(fl)
		if ctype is None or encoding is not None:
			ctype = "application/octet-stream"

		maintype, subtype = ctype.split("/", 1)

		if maintype == "text":
			fp = open(fl)
			# Note: we should handle calculating the charset
			attachment = MIMEText(fp.read(), _subtype=subtype)
			fp.close()
		elif maintype == "image":
			fp = open(fl, "rb")
			attachment = MIMEImage(fp.read(), _subtype=subtype)
			fp.close()
		elif maintype == "audio":
			fp = open(fl, "rb")
			attachment = MIMEAudio(fp.read(), _subtype=subtype)
			fp.close()
		else:
			fp = open(fl, "rb")
			attachment = MIMEBase(maintype, subtype)
			attachment.set_payload(fp.read())
			fp.close()
			encoders.encode_base64(attachment)
		attachment.add_header("Content-Disposition", "attachment", filename=fl)
		msg.attach(attachment)
		
		server = smtplib.SMTP('mail.maxionwheels.com',25)
		server.starttls()
		try:
			server.sendmail(emailfrom, emailto.split(';'), msg.as_string())
		except socket.error:
			print u"Nao foi possivel conectar ao servidor"
		else:
			print u"Email enviado com sucesso para " + emailto
	else:
		print u"O real não foi excedido em nenhum CC para a Conta: " + owner.Grupo[-3:]
	
		
ambiente = "aco"
conn(ambiente)
