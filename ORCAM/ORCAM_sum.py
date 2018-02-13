# encoding: utf-8
# encoding: iso-8859-1
# encoding: win-1252
## COMPILAR: pyinstaller.exe --onefile --icon=ORC_icon.ico ORCAM_sum.py
import pyodbc
import sys
import smtplib
import mimetypes
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
import shutil
import subprocess
import time
import os
from distutils.core import setup
from ConfigParser import SafeConfigParser
import glob
import socket
import os.path
from openpyxl import Workbook
from ctypes import *

wb = Workbook()
ws1 = wb.active
ws1.title = 'Aco'
ws2 = wb.create_sheet('Alum')

ws1['A1'] = 'Conta'
ws1['B1'] = 'Nome Conta'
ws1['C1'] = 'Valor (ac. do mes)'

ws2['A1'] = 'Conta'
ws2['B1'] = 'Nome Conta'
ws2['C1'] = 'Valor (ac. do mes)'


cdate = time.strftime("%Y%m")
ins_date = time.strftime("%Y%m%d")
emailfrom = "ORCAM_report@maxionwheels.com"
server = 'LIM-SQL12P1'
username = 'dsserver'
password = 'Maxion123@'
driver = '{SQL Server}' # Driver necessário para conecta ao banco de dados do MSSQL
port = '1433'
emailto = 'alex.bacalhau@maxionwheels.com;gilberto.ferreira@maxionwheels.com;marcos.degaspre@maxionwheels.com'
#emailto = 'rafael.oliveira@maxionwheels.com'

def conn(ambiente):
	if ambiente == "aco":
		database = 'PROD_ORCAM_LMS'
		cnn = pyodbc.connect('DRIVER='+driver+';PORT=port;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+
					';PWD='+password)
		cnn.setencoding(str, encoding='utf-8')
		cnn.setencoding(unicode, encoding='utf-8', ctype=pyodbc.SQL_CHAR)
		cursor = cnn.cursor()
		app(cnn, cursor)
	if ambiente == "alum":
		database = 'PROD_ORCAM_LMA'
		cnn = pyodbc.connect('DRIVER='+driver+';PORT=port;SERVER='+server+';PORT=1443;DATABASE='+database+';UID='+username+
					';PWD='+password)
		cnn.setencoding(str, encoding='utf-8')
		cnn.setencoding(unicode, encoding='utf-8', ctype=pyodbc.SQL_CHAR)
		cursor = cnn.cursor()
		ambAlum(cnn, cursor)

def ambAlum(cnn, cursor):
	row = cursor.execute("select * from Grupos where grupo = '0411' or grupo = '0412' or grupo = '0415' or grupo = '0416' or grupo = '0462' or grupo = '0469' or grupo = '0499' or grupo = '0500' or grupo = '0480' or grupo = '0481' or grupo = '0482' or grupo = '0496'")
	i = 2
	body4 = []
	for row in cursor.fetchall():
		for conta in cursor.execute("select sum(Vr_real) from orcam where anomes=? and grupo=? and Ccusto >= '0300' and Ccusto != '0510' and Ccusto <= '0864' ", cdate, row.Grupo):
			if conta[0] != None:
				
				value_real = str("%0.2f" % (conta[0])).replace('.',',')
				ws2.cell(row=i, column=1).value = row.Grupo
				ws2.cell(row=i, column=2).value = str(row.Descricao.encode('ascii', 'ignore'))				
				ws2.cell(row=i, column=3).value = value_real
				i = i + 1		 
		
		
def app(cnn, cursor):
	row = cursor.execute("select * from Grupos where grupo = '0411' or grupo = '0412' or grupo = '0415' or grupo = '0416' or grupo = '0462' or grupo = '0469' or grupo = '0499' or grupo = '0500' or grupo = '0480' or grupo = '0481' or grupo = '0482' or grupo = '0496'")
	i = 2
	body = []
	for row in cursor.fetchall():
		for conta in cursor.execute("select sum(Vr_real) from orcam where Grupo=? and anomes=? and Ccusto >= '0300' and Ccusto != '0510' and Ccusto <= '0601' or grupo=? and anomes=? and Ccusto >='0810' and Ccusto <= '0812' and Ccusto != '0510'", row.Grupo, cdate, row.Grupo, cdate):
			sub = u"ORÇAM: Valor Real por Conta"
			if conta[0] != None:
				value_real = str("%0.2f" % (conta[0])).replace('.',',')
				ws1.cell(row=i, column=1).value = row.Grupo
				ws1.cell(row=i, column=2).value = str(row.Descricao.encode('ascii', 'ignore'))
				ws1.cell(row=i, column=3).value = value_real
				i = i + 1
	emailfrom = "ORCAM_report@maxionwheels.com"
	body3 = u"""<html>
	<head>
	<font face="Calibri" size="3" color = "black">
	</head>
	<body>
		<b>AÇO:</b> soma do CC >= 0300 e <= 0600 e CC 0810 até 0812, com exceção do CC 0510
		<br>
		<b>ALUM:</b> soma do CC >= 0300 e <= 864 com exceção do CC 0510
	
	
	</body>
	</font>
	</html>"""
	conn('alum')
	fl = 'ORCAM_FIN_' + cdate + '.xlsx'
	print fl
	wb.save(fl)
	msglala = body3
	msg = MIMEMultipart()
	msg.attach(MIMEText(msglala, 'html', 'utf-8'))
	msg["From"] = emailfrom
	msg["Subject"] = sub
	msg["To"] = emailto
	ctype, encoding = mimetypes.guess_type(fl)
	if ctype is None or encoding is not None:
		ctype = "application/octet-stream"
	
	maintype, subtype = ctype.split("/", 1)
	
	if maintype == "text":
		fp = open(fl)
		# Note: we should handle calculating the charset
		attachment = MIMEText(fp.read(), _subtype=subtype)
		fp.close()
	elif maintype == "image":
		fp = open(fl, "rb")
		attachment = MIMEImage(fp.read(), _subtype=subtype)
		fp.close()
	elif maintype == "audio":
		fp = open(fl, "rb")
		attachment = MIMEAudio(fp.read(), _subtype=subtype)
		fp.close()
	else:
		fp = open(fl, "rb")
		attachment = MIMEBase(maintype, subtype)
		attachment.set_payload(fp.read())
		fp.close()
		encoders.encode_base64(attachment)
	attachment.add_header("Content-Disposition", "attachment", filename=fl)
	msg.attach(attachment)
	
	server = smtplib.SMTP('mail.maxionwheels.com',25)
	server.starttls()
	try:
		server.sendmail(emailfrom, emailto.split(';'), msg.as_string())
	except socket.error:
		print u"Nao foi possivel conectar ao servidor"
	else:
		print u"Email enviado com sucesso para " + emailto

		
ambiente = "aco"
conn(ambiente)
